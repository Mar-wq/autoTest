import json

import openpyxl
import os,sys
import regex as re
import allure
from common.parsefunc import parsefunc_obj

from common.loggerController import log
from common.requestion_assertion import requestion_assertion



BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
print(BASE_DIR)
sys.path.append(BASE_DIR)


class BaseFunction:
    def __init__(self):
        self.methods = [requestion_assertion.send_request, requestion_assertion.unified_assert]

    def transact(self,file_path):
        """
        目的是让模块的用例循环执行，读取出来的datacases是一个四维数组
        :return:
        """
        datacases = self.read_test_data_from_excel(file_path)
        for data in datacases:      #data为三维
            #runcase_obj.module_name = sheetnames[j]
            for data_i in data:          #i为二维数组
                yield data_i


    def read_test_data_from_excel(self,file_path):
        """
        读取excel内容，返回一个三维数组，用例之间为一个数组，步骤在一个数组
        :param file_path: excel文件路径
        :param sheetname: sheet名
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        #pdb.set_trace()
        sheetnames = wb.sheetnames
        #pdb.set_trace()

        list_datas = []

        for sheetname in sheetnames:

            # 初始化列表
            sheet = wb[sheetname]
            list_data = []
            current_list = None

            # 遍历每一行
            for row in sheet.iter_rows(min_row=2, values_only=True):
                #pdb.set_trace()
                if row[0] is not None:  # 如果第一列有内容
                    if current_list is None:
                        current_list = []
                        current_list.append([sheetname] + list(row))
                    else:
                        list_data.append(current_list)
                        current_list = []
                        current_list.append([sheetname] + list(row))
                else:  # 如果第一列为空
                    current_list.append(list(row))

            #如果最后一列不为空，则将最后一个列表添加到字典中
            if current_list is not None:
                list_data.append(current_list)

            list_datas.append(list_data)

        return list_datas

    def run_case(self,test_data):
        print()
        """
        用例分步执行
        :param test_data: 一个用例的二维list，最里层为步骤list
        :return:
        """
        #allure.dynamic.severity(self.define_cases_level(test_data[0][0]))
        allure.dynamic.title(test_data[0][2])
        log.info("============================{}模块：用例{}：{}========================".format(test_data[0][0],test_data[0][1],test_data[0][2]))
        previous_response = None
        for index, case in enumerate(test_data[1:], start=0):
            json_str = case[3]
            with allure.step(case[2]):
                # 统一参数处理
                params = self.parse_json(json_str)
                if previous_response is not None:
                    # 将上一轮的响应结果添加到参数中
                    params['unified_response'] = previous_response

                response = self.methods[index](params)
                previous_response = response



    #处理参数列表
    def process_list(self, input_list):
        '''
        :param input_list:待处理的参数列表
        :return:  processed_list: 去除''的字符串以及数值型字符串转成相应数值
        '''
        processed_list = []
        for item in input_list:
            # 去掉单引号包围的字符串中的单引号
            if item.startswith("'") and item.endswith("'"):
                item = item[1:-1]
            # 尝试将字符串转换为整数或浮点数
            elif item.isdigit():  # 纯数字字符串转换为整数
                item = int(item)
            elif item.replace('.', '', 1).isdigit():  # 检查是否为浮点数格式的字符串
                item = float(item)
            processed_list.append(item)
        return processed_list

    #函数参数用栈的方式获取，可以避免解析错误
    def split_params(self, params_str):
        params = []
        param = ""
        depth = 0

        for char in params_str:
            if char == ',' and depth == 0:
                params.append(param)
                param = ""
            else:
                param += char
                if char == '[':
                    depth += 1
                elif char == ']':
                    depth -= 1

        if param:
            params.append(param)

        return params
    #解析字符串当中存在多个$[]嵌套的情况
    def parse_nested_func(self, nested_func_str):
        pattern = re.compile(r'\$\[(.*)\]')
        match = pattern.match(nested_func_str)
        match_group = match.group(1)

        # 提取最外层函数名及其参数
        outer_func_pattern = re.compile(r"(\w+)\((.*)\)")
        match = outer_func_pattern.match(match_group)

        func_name = match.group(1)
        params_str = match.group(2)

        params_str = params_str.replace(' ', '')
        params_list = self.split_params(params_str)
        processed1_params_list = self.process_list(params_list)  #

        processed2_params_list = []
        for param in processed1_params_list:
            if isinstance(param, str) and '$' in param:
                param = self.parse_nested_func(param)
            processed2_params_list.append(param)

        params_tuple = tuple(processed2_params_list)
        func = getattr(parsefunc_obj, func_name)
        res = func(*params_tuple)

        return res



    #解析参数字典以及嵌套字典当中存在$[]的情况
    def parse_value(self, value):
        '''
        :param value:  字典当中的值对象
        :return:    如果是如果不含占位符，则直接原字符串，如果包含$[],则返回解析过后的结果
        '''

        # 匹配 $[content] 模式的正则表达式
        pattern = re.compile(r'\$\[(.*?)\]')

        if isinstance(value, str):
            match = pattern.match(value)
            if match:
                # 提取括号内的内容并做进一步的递归解析
                handle_str =  match.group(0)
                return self.parse_nested_func(handle_str)
            return value
        elif isinstance(value, dict):
            # 递归解析字典中的值
            return {k: self.parse_value(v) for k, v in value.items()}
        elif isinstance(value, list):
            # 递归解析列表中的值
            return [self.parse_value(item) for item in value]
        else:
            return value


    def parse_json(self, json_str):
        data = json.loads(json_str)
        params_dict = self.parse_value(data)
        return params_dict


basefunc = BaseFunction()